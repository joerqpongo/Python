import os
import glob
from openpyxl import load_workbook

# Rutas del archivo Excel y la carpeta de PDFs
ruta_excel = r"C:\_ComprobaFra\_FacturasJL.xlsx"
ruta_pdfs = r"C:\_ComprobaFra"

# Cargar el archivo Excel (openpyxl permite mantenerlo abierto en segundo plano)
wb = load_workbook(ruta_excel)
ws = wb.active  # Obtener la hoja activa (asumimos que es la primera)

# Función para renombrar archivos PDF
def renombrar_pdf(pdfJoseStyle, pdfMCStyle):
    """Renombra un archivo PDF del formato pdfJoseStyle al formato pdfMCStyle."""
    try:
        os.rename(f"{ruta_pdfs}/{pdfJoseStyle}.pdf", f"{ruta_pdfs}/{pdfMCStyle}.pdf")
        print(f"Archivo renombrado: {pdfJoseStyle}.pdf -> {pdfMCStyle}.pdf")
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo {pdfJoseStyle}.pdf")
    except Exception as e:
        print(f"Error al renombrar el archivo: {e}")

# Iterar sobre las filas del Excel (desde la 2 hasta la 1987)
for fila in range(2, 5):
    nombre_archivo = ws.cell(row=fila, column=19).value  # Obtener el nombre de archivo de la columna 19 (S)
    
    # Buscar archivos PDF que coincidan con el nombre (en la carpeta y subcarpetas)
    patron_busqueda = f"{ruta_pdfs}/**/{nombre_archivo}"
    resultados = glob.glob(patron_busqueda, recursive=True)

    # Si se encuentra al menos un PDF, marcar como "ENC" (Encontrado)
    if resultados:
        ws.cell(row=fila, column=20).value = "ENC"
        ws.cell(row=fila, column=21).value = nombre_archivo
        ws.cell(row=fila, column=22).value = ws.cell(row=fila,column=23).value  
        pdfJoseStyle=(nombre_archivo)   
        #print(pdfJoseStyle)
        pdfMCStyle=(ws.cell(row=fila,column=23).value)
        print(pdfJoseStyle," el formato de JR se va a cambiar por el de MC: ", pdfMCStyle)
        
        # Llamar a la función para renombrar el archivo
        #renombrar_pdf(pdfJoseStyle, pdfMCStyle)
        
    else:
        ws.cell(row=fila, column=20).value = "NA"  # Si no se encuentra, marcar como "NA"

# Guardar los cambios en el archivo Excel
wb.save(ruta_excel)

print("Proceso completado. Resultados guardados en:", ruta_excel)
