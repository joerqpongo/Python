import os
import glob
from openpyxl import load_workbook

# Rutas del archivo Excel y la carpeta de PDFs
ruta_excel = r"C:\_ComprobaFra\_FacturasJL.xlsx"
ruta_pdfs = r"C:\_ComprobaFra"
# Cargar el archivo Excel (openpyxl permite mantenerlo abierto en segundo plano)
wb = load_workbook(ruta_excel)
ws = wb.active  # Obtener la hoja activa (asumimos que es la primera)

# Función para renombrar archivos PDF
def renombrar_pdf(pdfJoseStyle, pdfMCStyle):
    """Renombra un archivo PDF del formato pdfJoseStyle al formato pdfMCStyle."""
    try:
        os.rename(f"{ruta_pdfs}/{pdfJoseStyle}", f"{ruta_pdfs}/{pdfMCStyle}")
        print(f"Archivo renombrado: {pdfJoseStyle} -> {pdfMCStyle}")
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo {pdfJoseStyle}")
    except Exception as e:
        print(f"Error al renombrar el archivo: {e}")

# Función para archivar archivos PDF
def archivar_pdf(pdfMCStyle, subcarpeta):
    """Archivar un archivo PDF a la ruta "C:\_ComprobaFra\" con la subcarpeta especificada."""
    try:
        ruta_archivo_original = f"{ruta_pdfs}/{pdfMCStyle}"
        ruta_archivo_archivado = f"{ruta_pdfs}/{subcarpeta}/{pdfMCStyle}"
        os.makedirs(f"{ruta_pdfs}/{subcarpeta}", exist_ok=True)  # Crear la subcarpeta si no existe
        os.rename(ruta_archivo_original, ruta_archivo_archivado)
        print(f"Archivo archivado: {pdfMCStyle} -> {ruta_archivo_archivado}")
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo {pdfMCStyle}")
    except Exception as e:
        print(f"Error al archivar el archivo: {e}")

# Iterar sobre las filas del Excel (desde la 2 hasta la 1987)
for fila in range(2, 5):
    nombre_archivo = ws.cell(row=fila, column=19).value  # Obtener el nombre de archivo de la columna 19 (S)
    
    # Buscar archivos PDF que coincidan con el nombre (en la carpeta y subcarpetas)
    patron_busqueda = f"{ruta_pdfs}/**/{nombre_archivo}"
    resultados = glob.glob(patron_busqueda, recursive=True)

    # Si se encuentra al menos un PDF, marcar como "ENC" (Encontrado)
    if resultados:
        ws.cell(row=fila, column=20).value = "ENC"
        ws.cell(row=fila, column=21).value = nombre_archivo
        ws.cell(row=fila, column=22).value = ws.cell(row=fila,column=23).value  
        pdfJoseStyle=(nombre_archivo)   
        #print(pdfJoseStyle)
        pdfMCStyle=(ws.cell(row=fila,column=23).value)
        print(pdfJoseStyle," el formato de JR se va a cambiar por el de MC: ", pdfMCStyle)
        
        # Llamar a la función para renombrar el archivo
        #renombrar_pdf(pdfJoseStyle, pdfMCStyle)

        # Obtener la subcarpeta de la columna 17
        subcarpeta = ws.cell(row=fila, column=24).value
        print(subcarpeta)

        # Llamar a la función para archivar el archivo
        archivar_pdf(pdfJoseStyle, subcarpeta)
        
        # Llamar a la función para renombrar el archivo
        renombrar_pdf(pdfJoseStyle, pdfMCStyle)

        
    else:
        ws.cell(row=fila, column=20).value = "NA"  # Si no se encuentra, marcar como "NA"

# Guardar los cambios en el archivo Excel
wb.save(ruta_excel)

print("Proceso completado. Resultados guardados en:", ruta_excel)
